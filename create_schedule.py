from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 평일만 계산하는 함수
def get_weekdays(start_date, end_date):
    weekdays = []
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # 0-4는 월-금
            weekdays.append(current)
        current += timedelta(days=1)
    return weekdays

# 프로젝트 기간 설정
start_date = datetime(2026, 1, 15)
end_date = datetime(2026, 2, 12)
weekdays = get_weekdays(start_date, end_date)

print(f"총 평일 수: {len(weekdays)}일")

# 엑셀 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "일정계획표"

# 헤더 작성
ws['A1'] = '추진내용'
ws.merge_cells('A1:A2')

# 날짜 헤더 작성
date_header = f"{start_date.month}/{start_date.day} ~ {end_date.month}/{end_date.day}"
end_col = get_column_letter(len(weekdays) + 1)
ws.merge_cells(f'B1:{end_col}1')
ws['B1'] = date_header
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 날짜 열 헤더
for idx, day in enumerate(weekdays, start=2):
    col = get_column_letter(idx)
    ws[f'{col}2'] = f"{day.month}/{day.day}"

# 일정 인덱스 계산
req_start_idx = 0  # 1/15
req_end_idx = 2    # 1/17 (3일)
design_start_idx = 1  # 1/16 (2일차)
design_end_idx = 3    # 1/19 (4일차)
dev_start_idx = 3     # 1/19 (4일차)
dev_end_idx = len(weekdays) - 4  # 2/10 (마지막일-3일)
test_start_idx = len(weekdays) - 3  # 2/10
test_end_idx = len(weekdays) - 1    # 2/12

# 행 번호 추적
row = 3

# 활동 작성 함수
def write_activity(name, level=0, start_idx=None, end_idx=None, fill_color='D3D3D3'):
    global row
    ws[f'A{row}'] = '  ' * level + name
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    if start_idx is not None and end_idx is not None:
        for idx in range(start_idx, end_idx + 1):
            col = get_column_letter(idx + 2)
            cell = ws[f'{col}{row}']
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    row += 1
    return row - 1

# 추진내용 작성
# 요구분석
write_activity('요구분석', 0, req_start_idx, req_end_idx)
write_activity('기획', 1)
write_activity('기획안 작성', 2, req_start_idx, req_start_idx)
write_activity('개발환경 설정', 2, req_start_idx, req_start_idx)
write_activity('역할분담', 2, req_start_idx, req_start_idx)
write_activity('분석', 1)
write_activity('요구사항 수집/정리', 2, req_start_idx, req_end_idx)
write_activity('시스템 아키텍쳐 정의', 2, req_start_idx, req_end_idx)

# 설계
write_activity('설계', 0, design_start_idx, design_end_idx)
write_activity('DB설계', 1)
write_activity('ERD', 2, design_start_idx, design_end_idx)
write_activity('화면설계', 1)
write_activity('스토리보드', 2, design_start_idx, design_end_idx)

# 개발
write_activity('개발', 0, dev_start_idx, dev_end_idx)
write_activity('프론트엔드', 1)
write_activity('동적 웹 페이지 개발', 2, dev_start_idx, dev_end_idx)
write_activity('사용자 입력/결과 페이지 구현', 2, dev_start_idx, dev_end_idx)
write_activity('백엔드', 1)
write_activity('API개발', 2, dev_start_idx, dev_end_idx)
write_activity('SSO개발', 2, dev_start_idx, dev_end_idx)
write_activity('AWS연동', 2, dev_start_idx, dev_end_idx)
write_activity('지출데이터', 2, dev_start_idx, dev_end_idx)
write_activity('소비패턴분석', 2, dev_start_idx, dev_end_idx)
write_activity('크롤링 카드데이터', 2, dev_start_idx, dev_end_idx)
write_activity('데이터 정제', 1)
write_activity('데이터 수집/전처리', 2, dev_start_idx, dev_end_idx)
write_activity('머신러닝', 1)
write_activity('모델학습 및 성능 개선', 2, dev_start_idx, dev_end_idx)
write_activity('LLM 카드추천', 1)
write_activity('LLM통합 및 튜닝', 2, dev_start_idx, dev_end_idx)
write_activity('LLM 소비 분석', 1)
write_activity('LLM통합 및 튜닝', 2, dev_start_idx, dev_end_idx)

# 테스트
write_activity('테스트', 0, test_start_idx, test_end_idx)
write_activity('단위테스트', 1, test_start_idx, test_end_idx)
write_activity('통합테스트', 1, test_start_idx, test_end_idx)

# 최종보고서 작성 및 수정보완
write_activity('최종보고서 작성 및 수정보완', 0, test_start_idx, test_end_idx, '90EE90')

# 프로젝트 수행착수/중간/최종보고
write_activity('프로젝트 수행착수/중간/최종보고', 0, 0, len(weekdays) - 1)

# 열 너비 조정
ws.column_dimensions['A'].width = 50
for idx in range(2, len(weekdays) + 2):
    col = get_column_letter(idx)
    ws.column_dimensions[col].width = 8

# 스타일 적용
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 헤더 스타일
ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].border = border
ws['A2'].fill = header_fill
ws['A2'].font = header_font
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws['A2'].border = border

ws['B1'].fill = header_fill
ws['B1'].font = header_font
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].border = border

for idx in range(2, len(weekdays) + 2):
    col = get_column_letter(idx)
    ws[f'{col}2'].fill = header_fill
    ws[f'{col}2'].font = header_font
    ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col}2'].border = border

# 모든 셀에 테두리 적용
for row_num in range(1, row):
    for col_num in range(1, len(weekdays) + 2):
        col = get_column_letter(col_num)
        cell = ws[f'{col}{row_num}']
        if cell.border.left.style is None:
            cell.border = border

# 파일 저장
output_file = r'd:\VScode\smartBudget\프로젝트_일정계획표.xlsx'
wb.save(output_file)
print(f'엑셀 파일이 생성되었습니다: {output_file}')